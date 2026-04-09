"""
ViewSets (views) for Certificate and Delivery APIs
"""
from rest_framework import viewsets, status, permissions, serializers
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authentication import SessionAuthentication
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import openpyxl
import csv
import zipfile
import io
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image
import logging

logger = logging.getLogger(__name__)

from certificados.models import Certificate, Template
from deliveries.models import DeliveryLog
from students.models import Student
from events.models import Event, Enrollment, EventCategory
from core.permissions import AnonymousAccessLimit, AllowAnyWithLimit, IsAdminUser, get_client_ip
from .serializers import (
    CertificateListSerializer,
    CertificateDetailSerializer,
    CertificateCreateSerializer,
    CertificateGenerateSerializer,
    CertificateDeliverSerializer,
    DeliveryLogSerializer,
)


# ─────────────────────────────────────────
# PERMISOS PERSONALIZADOS
# ─────────────────────────────────────────

class IsAdminUserOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user and request.user.is_staff


class IsCertificateOwnerOrAdmin(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_staff


# ─────────────────────────────────────────
# SERIALIZERS INLINE
# ─────────────────────────────────────────

class TemplateSerializer(serializers.ModelSerializer):
    class Meta:
        model = Template
        fields = [
            'id', 'name', 'category', 'background_url', 'preview_url',
            'x_coord', 'y_coord', 'font_size', 'font_color', 'font_family',
        ]
        read_only_fields = ['id', 'created_by', 'created_at', 'updated_at']


# ─────────────────────────────────────────
# VIEWSETS
# ─────────────────────────────────────────

class CertificateViewSet(viewsets.ModelViewSet):
    queryset = Certificate.objects.all()

    def get_queryset(self):
        if self.request.user.is_staff:
            return Certificate.objects.all().select_related(
                'student', 'event', 'template', 'generated_by'
            )
        return Certificate.objects.filter(
            student__user=self.request.user
        ).select_related('student', 'event', 'template', 'generated_by')

    def get_serializer_class(self):
        if self.action == 'create':
            return CertificateCreateSerializer
        elif self.action == 'generate':
            return CertificateGenerateSerializer
        elif self.action == 'deliver':
            return CertificateDeliverSerializer
        elif self.action == 'list':
            return CertificateListSerializer
        return CertificateDetailSerializer

    def get_permissions(self):
        if self.action == 'verify':
            self.permission_classes = [permissions.AllowAny]
        elif self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
        elif self.action in ['generate', 'deliver']:
            self.permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
        else:
            self.permission_classes = [permissions.IsAuthenticated]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        certificate = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            template_id = serializer.validated_data.get('template_id')
            if template_id:
                template = Template.objects.get(id=template_id)
                certificate.template = template
            certificate.generate()
            return Response({
                'status': 'success',
                'message': 'Certificate generated successfully',
                'certificate': CertificateDetailSerializer(certificate).data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': f'Failed to generate certificate: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def deliver(self, request, pk=None):
        certificate = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        method = serializer.validated_data['method']
        recipient = serializer.validated_data.get('recipient')
        try:
            with transaction.atomic():
                certificate.deliver(
                    method=method,
                    recipient=recipient,
                    sent_by=request.user
                )
                return Response({
                    'status': 'success',
                    'message': f'Certificate delivered via {method}',
                    'delivery_log': DeliveryLogSerializer(
                        certificate.last_delivery_attempt
                    ).data,
                    'certificate': CertificateDetailSerializer(certificate).data
                }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': f'Failed to deliver certificate: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        certificate = self.get_object()
        delivery_logs = certificate.get_delivery_history()
        return Response({
            'certificate_id': str(certificate.id),
            'total_attempts': delivery_logs.count(),
            'deliveries': DeliveryLogSerializer(delivery_logs, many=True).data
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny])
    def verify(self, request):
        code = request.query_params.get('code')
        if not code:
            return Response({
                'status': 'error',
                'message': 'Verification code is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        try:
            certificate = Certificate.objects.get(verification_code=code)
            if certificate.is_expired():
                return Response({
                    'status': 'error',
                    'message': 'Certificate has expired',
                    'certificate': CertificateDetailSerializer(certificate).data
                }, status=status.HTTP_410_GONE)
            return Response({
                'status': 'success',
                'message': 'Certificate verified successfully',
                'certificate': CertificateDetailSerializer(certificate).data
            }, status=status.HTTP_200_OK)
        except Certificate.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Certificate not found'
            }, status=status.HTTP_404_NOT_FOUND)


class DeliveryLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DeliveryLog.objects.all().select_related(
        'certificate', 'sent_by'
    ).order_by('-sent_at')
    serializer_class = DeliveryLogSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def get_queryset(self):
        queryset = super().get_queryset()
        cert_id = self.request.query_params.get('certificate_id')
        if cert_id:
            queryset = queryset.filter(certificate__id=cert_id)
        return queryset


# ─────────────────────────────────────────
# ENDPOINTS DE TEMPLATES (con autenticación)
# ─────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def crear_template_con_imagen(request):
    """
    Crear template con imagen y coordenadas.
    Requiere autenticación y modo admin activo.
    POST /api/templates/crear/
    Body: multipart/form-data con:
        - name: nombre del template
        - imagen: archivo de imagen
        - x_coord: coordenada X
        - y_coord: coordenada Y
        - font_size: tamaño de fuente
        - font_color: color en hex (default #000000)
        - event_id: ID del evento (opcional)
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        name = request.data.get('name')
        imagen = request.FILES.get('imagen')
        x_coord = float(request.data.get('x_coord', 200))
        y_coord = float(request.data.get('y_coord', 300))
        font_size = int(request.data.get('font_size', 24))
        font_color = request.data.get('font_color', '#000000')
        event_id = request.data.get('event_id')

        if not name or not imagen:
            return Response({
                'error': 'Se requiere nombre e imagen'
            }, status=status.HTTP_400_BAD_REQUEST)

        file_path = default_storage.save(
            f'templates/{imagen.name}',
            ContentFile(imagen.read())
        )

        template = Template.objects.create(
            created_by=request.user,
            name=name,
            background_url=default_storage.url(file_path),
            x_coord=x_coord,
            y_coord=y_coord,
            font_size=font_size,
            font_color=font_color,
            is_active=True
        )

        if event_id:
            try:
                event = Event.objects.get(id=event_id)
                event.template = template
                event.save()
            except Event.DoesNotExist:
                pass

        return Response({
            'success': True,
            'template': TemplateSerializer(template).data,
            'message': 'Template creado exitosamente'
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def listar_templates(request):
    """
    Listar templates del usuario autenticado.
    GET /api/templates/
    """
    templates = Template.objects.filter(
        created_by=request.user,
        is_active=True
    ).order_by('-created_at')

    return Response(TemplateSerializer(templates, many=True).data)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def actualizar_template(request, template_id):
    """
    Actualizar coordenadas y estilo de un template.
    Requiere autenticación y modo admin activo.
    PUT /api/templates/<template_id>/
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        template = Template.objects.get(id=template_id, created_by=request.user)

        if 'x_coord' in request.data:
            template.x_coord = float(request.data['x_coord'])
        if 'y_coord' in request.data:
            template.y_coord = float(request.data['y_coord'])
        if 'font_size' in request.data:
            template.font_size = int(request.data['font_size'])
        if 'font_color' in request.data:
            template.font_color = request.data['font_color']
        if 'font_family' in request.data:
            template.font_family = request.data['font_family']

        template.save()

        return Response({
            'success': True,
            'template': TemplateSerializer(template).data
        })

    except Template.DoesNotExist:
        return Response({
            'error': 'Template no encontrado'
        }, status=status.HTTP_404_NOT_FOUND)


# ─────────────────────────────────────────
# ENDPOINTS CUSTOM SIN AUTENTICACIÓN
# ─────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_students_and_create_certs(request):
    """
    Importa estudiantes desde Excel y crea certificados.
    Requiere autenticación y modo admin activo.
    POST /api/import/
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )
    """
    Importa estudiantes desde Excel y crea certificados.
    POST /api/import/
    Body: multipart/form-data con 'file' y 'event_id'
    """
    archivo = request.FILES.get('file')
    event_id = request.data.get('event_id')

    if not archivo or not event_id:
        return Response({'error': 'Se requiere file y event_id'}, status=400)

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Evento no encontrado'}, status=404)

    estudiantes = []
    nombre_archivo = archivo.name.lower()

    if nombre_archivo.endswith('.csv'):
        decoded = archivo.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)
        for row in reader:
            estudiantes.append(row)
    else:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            estudiantes.append(dict(zip(headers, row)))

    creados = 0
    certs_creados = 0
    errores = []

    for data in estudiantes:
        try:
            email = str(data.get('email', '')).strip()
            if not email or email == 'None':
                continue

            student, _ = Student.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': str(data.get('first_name', '')).strip(),
                    'last_name':  str(data.get('last_name', '')).strip(),
                    'document_id': str(data.get('document_id', email)).strip(),
                    'phone': str(data.get('phone', '')).strip(),
                }
            )
            creados += 1

            Enrollment.objects.get_or_create(student=student, event=event)

            cert, created = Certificate.objects.get_or_create(
                student=student,
                event=event,
                defaults={'status': 'pending'}
            )
            if created:
                certs_creados += 1

        except Exception as e:
            errores.append(str(e))

    return Response({
        'ok': True,
        'estudiantes_procesados': creados,
        'certificados_creados': certs_creados,
        'errores': errores[:5]
    })


@api_view(['GET'])
@permission_classes([AllowAnyWithLimit])
def listar_eventos(request):
    """
    GET /api/eventos/
    Lista eventos activos. Accesible sin login (límite 30 personas) o con login.
    """
    eventos = Event.objects.filter(is_active=True).values(
        'id', 'name', 'category', 'event_date'
    )
    return Response(list(eventos))


@api_view(['GET'])
@permission_classes([AllowAnyWithLimit])
def listar_certificados(request):
    """
    GET /api/certs/?event_id=1
    Lista certificados de un evento. Accesible sin login (límite 30) o con login.
    """
    event_id = request.query_params.get('event_id')
    if not event_id:
        return Response({'error': 'Se requiere event_id'}, status=400)

    certs = Certificate.objects.filter(
        event_id=event_id
    ).select_related('student').order_by('student__first_name')

    data = [{
        'id': c.id,
        'status': c.status,
        'student_name': c.student.first_name + ' ' + c.student.last_name,
        'student_email': c.student.email,
    } for c in certs]

    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_todos(request):
    """
    Genera todos los certificados pendientes de un evento.
    Requiere autenticación y modo admin activo.
    POST /api/generar/ body: {event_id: 1}
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )

    event_id = request.data.get('event_id')
    if not event_id:
        return Response({'error': 'Se requiere event_id'}, status=400)

    certs = Certificate.objects.filter(event_id=event_id, status='pending')
    ok = 0
    errores = []

    for cert in certs:
        try:
            cert.generate()
            ok += 1
        except Exception as e:
            errores.append(f'{cert.student.email}: {str(e)}')

    return Response({'generados': ok, 'errores': errores[:5]})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def enviar_todos(request):
    """
    Envía todos los certificados generados de un evento.
    Requiere autenticación y modo admin activo.
    POST /api/enviar/ body: {event_id: 1}
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )

    event_id = request.data.get('event_id')
    if not event_id:
        return Response({'error': 'Se requiere event_id'}, status=400)

    certs = Certificate.objects.filter(event_id=event_id, status='generated')
    ok = 0
    errores = []

    for cert in certs:
        try:
            cert.deliver(method='email', recipient=cert.student.email)
            ok += 1
        except Exception as e:
            errores.append(f'{cert.student.email}: {str(e)}')

    return Response({'enviados': ok, 'errores': errores[:5]})


@api_view(['POST'])
@permission_classes([AllowAnyWithLimit])
@parser_classes([MultiPartParser, FormParser])
def generar_certificados_publico(request):
    """
    Modo Público: Genera certificados sin autenticación (límite 30 personas).
    POST /api/publico/generar/
    Body: multipart/form-data con:
        - imagen: archivo de imagen (PNG/JPG)
        - excel: archivo Excel con estudiantes
        - x_coord: coordenada X para el nombre (float)
        - y_coord: coordenada Y para el nombre (float)
        - font_size: tamaño de fuente (int, default 24)
    Returns: ZIP con todos los PDFs generados
    """
    # Verificar si se alcanzó el límite (el permission class ya verifica, pero por si acaso)
    from core.models import AnonymousAccess
    session_key = request.session.session_key
    if not request.user.is_authenticated:
        if not AnonymousAccess.can_access(session_key):
            return Response({
                'error': 'Límite alcanzado',
                'message': 'Se ha alcanzado el límite de 30 usuarios. Inicia sesión para continuar.'
            }, status=status.HTTP_403_FORBIDDEN)
    try:
        imagen_plantilla = request.FILES.get('imagen')
        archivo_excel = request.FILES.get('excel')
        x_coord = float(request.data.get('x_coord', 200))
        y_coord = float(request.data.get('y_coord', 300))
        font_size = int(request.data.get('font_size', 24))

        if not imagen_plantilla or not archivo_excel:
            return Response({
                'error': 'Debes subir la imagen plantilla y el archivo Excel'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(archivo_excel)
        except Exception as e:
            return Response({
                'error': f'Error al leer el Excel: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        required_columns = ['first_name', 'last_name', 'email']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return Response({
                'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}',
                'required_columns': required_columns
            }, status=status.HTTP_400_BAD_REQUEST)

        zip_buffer = io.BytesIO()
        generated_count = 0
        errors = []

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for idx, row in df.iterrows():
                try:
                    nombre = str(row['first_name']).strip()
                    apellido = str(row['last_name']).strip()
                    nombre_completo = f"{nombre} {apellido}"
                    document_id = str(row.get('document_id', f'estudiante_{idx}')).strip()

                    pdf_buffer = io.BytesIO()

                    img = Image.open(imagen_plantilla)
                    img_width, img_height = img.size

                    c = canvas.Canvas(pdf_buffer, pagesize=(img_width, img_height))
                    c.drawImage(
                        ImageReader(img),
                        0, 0,
                        width=img_width,
                        height=img_height,
                        preserveAspectRatio=True,
                        mask='auto'
                    )

                    c.setFont("Helvetica-Bold", font_size)
                    c.setFillColorRGB(0, 0, 0)

                    palabras = nombre_completo.split()
                    if len(palabras) >= 2:
                        c.drawString(x_coord, y_coord + (font_size * 0.8), palabras[0])
                        c.drawString(x_coord, y_coord, ' '.join(palabras[1:]))
                    else:
                        c.drawString(x_coord, y_coord, nombre_completo)

                    c.save()

                    filename = f"{nombre}_{apellido}_{document_id}.pdf"
                    filename = filename.replace(' ', '_').replace('/', '_').replace('\\', '_')

                    zip_file.writestr(filename, pdf_buffer.getvalue())
                    generated_count += 1

                except Exception as e:
                    errors.append(f"{nombre_completo}: {str(e)}")
                    continue

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="certificados_generados.zip"'
        return response

    except Exception as e:
        return Response({
            'error': f'Error general: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─────────────────────────────────────────
# ENDPOINTS DE PERFIL Y MODO ADMIN
# ─────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mi_perfil(request):
    """
    Obtiene el perfil del usuario autenticado con sus datos,
    eventos donde participó y certificados obtenidos.
    GET /api/perfil/
    """
    user = request.user

    # Datos básicos del usuario
    data = {
        'id': user.id,
        'email': user.email,
        'full_name': user.full_name,
        'role': user.role,
        'is_admin': user.is_admin,
        'admin_mode_enabled': user.admin_mode_enabled,
        'is_admin_mode': user.is_admin_mode,
        'created_at': user.created_at,
        'email_app_password_configured': bool(user.email_app_password),
    }

    # Si tiene perfil de estudiante, agregar datos
    if hasattr(user, 'student_profile') and user.student_profile:
        student = user.student_profile
        data['student_profile'] = {
            'id': student.id,
            'document_id': student.document_id,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'email': student.email,
            'phone': student.phone,
        }

        # Eventos donde está inscrito
        enrollments = student.enrollments.select_related('event').all()
        data['enrollments'] = [{
            'event_id': e.event.id,
            'event_name': e.event.name,
            'event_date': e.event.event_date,
            'attendance': e.attendance,
            'enrolled_at': e.enrolled_at,
        } for e in enrollments]

        # Certificados obtenidos
        certs = student.certificates.select_related('event').all()
        data['certificates'] = [{
            'id': c.id,
            'event_name': c.event.name,
            'status': c.status,
            'verification_code': c.verification_code,
            'issued_at': c.issued_at,
            'pdf_url': c.pdf_url,
        } for c in certs]
    else:
        data['student_profile'] = None
        data['enrollments'] = []
        data['certificates'] = []

    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def toggle_admin_mode(request):
    """
    Activa/desactiva el modo admin para usuarios con rol admin.
    POST /api/perfil/toggle-admin/
    """
    user = request.user

    if not user.is_admin:
        return Response({
            'error': 'Solo los usuarios admin pueden activar el modo administrador'
        }, status=status.HTTP_403_FORBIDDEN)

    new_mode = user.toggle_admin_mode()

    return Response({
        'success': True,
        'admin_mode_enabled': new_mode,
        'message': 'Modo admin activado' if new_mode else 'Modo admin desactivado'
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def guardar_app_password(request):
    """
    Guarda el App Password de Gmail del usuario para enviar invitaciones.
    POST /api/perfil/guardar-app-password/
    Body: { app_password: "xxxx xxxx xxxx xxxx" }
    """
    user = request.user
    app_password = request.data.get('app_password', '').strip().replace(' ', '')
    
    if not app_password:
        return Response({
            'error': 'El App Password no puede estar vacío'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if len(app_password) != 16:
        return Response({
            'error': 'El App Password debe tener 16 caracteres'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    user.email_app_password = app_password
    user.save(update_fields=['email_app_password'])
    
    logger.info(f"App Password configurado para {user.email}")
    
    return Response({
        'success': True,
        'message': 'App Password guardado correctamente',
        'email_app_password_configured': True
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_email_config(request):
    """
    Prueba la configuración de email enviando un test.
    POST /api/perfil/test-email/
    """
    user = request.user
    
    if not user.email_app_password:
        return Response({
            'success': False,
            'error': 'No tienes App Password configurado'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from django.core.mail import EmailMessage, SMTPException
        from django.core.mail.backends.smtp import EmailBackend
        
        test_email = EmailMessage(
            subject='✅ Test - CertifyPro',
            body=f'''Hola {user.full_name},

Este es un email de prueba para verificar tu configuración.

Si recibes este mensaje, tu App Password está configurado correctamente.

--
CertifyPro''',
            from_email=user.email,
            to=[user.email]
        )
        
        connection = EmailBackend(
            host='smtp.gmail.com',
            port=587,
            username=user.email,
            password=user.email_app_password,
            use_tls=True,
            fail_silently=False
        )
        
        test_email.connection = connection
        test_email.send(fail_silently=False)
        connection.close()
        
        return Response({
            'success': True,
            'message': f'Email de prueba enviado a {user.email}. Revisa tu bandeja de entrada.'
        })
        
    except SMTPException as e:
        error_msg = str(e)
        logger.error(f"Error SMTP en test: {error_msg}")
        
        if 'authentication' in error_msg.lower():
            return Response({
                'success': False,
                'error': 'App Password incorrecto. Verifica que el código tenga 16 caracteres sin espacios.'
            }, status=status.HTTP_400_BAD_REQUEST)
        elif 'could not connect' in error_msg.lower() or 'connection' in error_msg.lower():
            return Response({
                'success': False,
                'error': 'No se pudo conectar al servidor de Gmail. Verifica tu conexión.'
            }, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({
                'success': False,
                'error': f'Error: {error_msg}'
            }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ─────────────────────────────────────────
# ENDPOINTS PARA ESTUDIANTES (MIS EVENTOS/CERTIFICADOS)
# ─────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_eventos(request):
    """
    Obtiene los eventos donde el usuario está inscrito.
    GET /api/mis-eventos/
    """
    user = request.user

    if not hasattr(user, 'student_profile') or not user.student_profile:
        return Response({
            'events': [],
            'message': 'No tienes un perfil de estudiante vinculado'
        })

    student = user.student_profile
    enrollments = student.enrollments.select_related('event', 'event__category').all()

    events_data = []
    for enrollment in enrollments:
        event = enrollment.event
        events_data.append({
            'event_id': event.id,
            'name': event.name,
            'description': event.description,
            'category': event.category.name if event.category else None,
            'event_date': event.event_date,
            'location': event.location,
            'attendance': enrollment.attendance,
            'enrolled_at': enrollment.enrolled_at,
        })

    return Response({
        'student': {
            'id': student.id,
            'full_name': student.full_name,
        },
        'events': events_data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_certificados(request):
    """
    Obtiene los certificados del usuario.
    GET /api/mis-certificados/
    """
    user = request.user

    if not hasattr(user, 'student_profile') or not user.student_profile:
        return Response({
            'certificates': [],
            'message': 'No tienes un perfil de estudiante vinculado'
        })

    student = user.student_profile
    certs = student.certificates.select_related('event', 'template').all()

    certs_data = []
    for cert in certs:
        certs_data.append({
            'id': cert.id,
            'event_name': cert.event.name,
            'event_date': cert.event.event_date,
            'status': cert.status,
            'verification_code': cert.verification_code,
            'pdf_url': cert.pdf_url,
            'issued_at': cert.issued_at,
            'template_name': cert.template.name if cert.template else None,
        })

    return Response({
        'student': {
            'id': student.id,
            'full_name': student.full_name,
        },
        'certificates': certs_data
    })


# ─────────────────────────────────────────
# ENDPOINTS DE ADMIN - CRUD EVENTOS
# ─────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_evento(request):
    """
    Crea un nuevo evento.
    Requiere autenticación y modo admin activo.
    POST /api/eventos/crear/
    Body: {
        name: string
        description: string (opcional)
        event_date: YYYY-MM-DD
        end_date: YYYY-MM-DD (opcional)
        duration_hours: int (opcional)
        location: string (opcional)
        category_id: int (opcional)
    }
    """
    # Verificar modo admin
    if not request.user.is_admin_mode:
        return Response(
            {'error': 'Solo disponible en modo administrador'},
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        name = request.data.get('name')
        event_date = request.data.get('event_date')

        if not name or not event_date:
            return Response(
                {'error': 'Se requiere name y event_date'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear el evento
        event = Event.objects.create(
            name=name,
            description=request.data.get('description', ''),
            event_date=event_date,
            end_date=request.data.get('end_date'),
            duration_hours=request.data.get('duration_hours'),
            location=request.data.get('location', ''),
            created_by=request.user,
            status='active',
            is_active=True
        )

        # Asignar categoría si se proporcionó
        category_id = request.data.get('category_id')
        if category_id:
            try:
                category = EventCategory.objects.get(id=category_id)
                event.category = category
                event.save()
            except EventCategory.DoesNotExist:
                pass  # Evento se crea sin categoría

        return Response({
            'success': True,
            'event': {
                'id': event.id,
                'name': event.name,
                'event_date': event.event_date,
                'status': event.status,
            },
            'message': 'Evento creado exitosamente'
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({
            'error': f'Error al crear evento: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─────────────────────────────────────────
# ENDPOINT DE ESTADO DE ACCESO PÚBLICO
# ─────────────────────────────────────────

@api_view(['GET'])
@permission_classes([AllowAny])
def estado_acceso_publico(request):
    """
    Consulta el estado del acceso público (límite de 30).
    GET /api/publico/estado/
    """
    from core.models import AnonymousAccess

    total_accesos = AnonymousAccess.get_active_count()
    cupos_disponibles = max(0, 30 - total_accesos)

    session_key = request.session.session_key
    if not session_key:
        request.session.create()
        session_key = request.session.session_key

    tiene_acceso = AnonymousAccess.can_access(session_key)

    return Response({
        'total_accesos': total_accesos,
        'cupo_maximo': 30,
        'cupos_disponibles': cupos_disponibles,
        'tiene_acceso': tiene_acceso or request.user.is_authenticated,
        'requiere_login': cupos_disponibles <= 0 and not request.user.is_authenticated,
    })


# ─────────────────────────────────────────
# AUTENTICACIÓN GOOGLE
# ─────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
def google_auth(request):
    """
    Recibe el token de Google y crea/retorna el usuario.
    POST /api/auth/google/
    Body: { credential: "google_token_jwt" }
    """
    import base64
    import json
    
    credential = request.data.get('credential')
    if not credential:
        return Response(
            {'error': 'Token de Google requerido'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Decodificar el JWT de Google (payload)
        # El JWT tiene formato: header.payload.signature
        parts = credential.split('.')
        if len(parts) != 3:
            return Response(
                {'error': 'Token de Google inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Decodificar el payload (parte del medio)
        payload = parts[1]
        # Añadir padding si es necesario
        padding = 4 - len(payload) % 4
        if padding != 4:
            payload += '=' * padding
        
        decoded = base64.urlsafe_b64decode(payload)
        user_data = json.loads(decoded)
        
        email = user_data.get('email')
        name = user_data.get('name', '')
        picture = user_data.get('picture', '')
        google_id = user_data.get('sub', '')
        
        if not email:
            return Response(
                {'error': 'Email no encontrado en token'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Buscar o crear usuario
        from users.models import User
        user, created = User.objects.get_or_create(
            email=email,
            defaults={
                'full_name': name,
                'is_active': True,
            }
        )
        
        if created:
            logger.info(f"Usuario creado desde Google: {email}")
        
        # Preparar respuesta
        return Response({
            'success': True,
            'user': {
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'is_admin': user.is_admin,
                'admin_mode_enabled': user.admin_mode_enabled,
                'is_admin_mode': user.is_admin_mode,
                'picture': picture,
            },
            'token': credential,  # Devolver el mismo token de Google
        })
        
    except Exception as e:
        logger.error(f"Error en autenticación Google: {str(e)}")
        return Response(
            {'error': f'Error al procesar token: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


# ─────────────────────────────────────────
# VISTAS PARA EVENTOS - GENERAR CERTIFICADOS
# ─────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def generar_certificados_evento(request, event_id):
    """
    Genera certificados para un evento desde Excel.
    Permite hasta 30 certificados sin autenticación.
    POST /api/events/{event_id}/generar/
    """
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Evento no encontrado'}, status=404)
    
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'Se requiere archivo Excel'}, status=400)
    
    try:
        if excel_file.name.endswith('.csv'):
            df = pd.read_csv(excel_file)
        else:
            df = pd.read_excel(excel_file)
        
        df.columns = df.columns.str.lower().str.strip()
        
        created = 0
        skipped = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                email = str(row.get('email', '')).strip().lower()
                nombre = str(row.get('nombre', '')).strip()
                
                if not email or '@' not in email:
                    errors.append(f"Fila {idx + 2}: Email inválido")
                    skipped += 1
                    continue
                
                # Crear estudiante si no existe
                student, _ = Student.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': nombre.split()[0] if nombre else 'Estudiante',
                        'last_name': ' '.join(nombre.split()[1:]) if nombre else '',
                        'document_id': email.split('@')[0],
                    }
                )
                
                # Crear certificado
                cert, cert_created = Certificate.objects.get_or_create(
                    event=event,
                    student=student,
                    defaults={'status': 'pending'}
                )
                
                if cert_created:
                    created += 1
                else:
                    skipped += 1
                    
            except Exception as e:
                errors.append(f"Fila {idx + 2}: {str(e)}")
                skipped += 1
        
        return Response({
            'success': True,
            'generated': created,
            'skipped': skipped,
            'errors': errors[:10]
        })
        
    except Exception as e:
        logger.error(f"Error generando certificados: {str(e)}")
        return Response({'error': str(e)}, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
def enviar_certificados_evento(request, event_id):
    """
    Envía certificados generados por email con PDF adjunto.
    POST /api/events/{event_id}/enviar-certificados/
    """
    from smtplib import SMTPException
    from django.core.mail import EmailMessage
    from django.core.mail.backends.smtp import EmailBackend
    
    user = request.user if request.user.is_authenticated else None
    user_email = user.email if user and user.is_authenticated else None
    user_app_password = user.email_app_password if user and user.is_authenticated else None
    
    logger.info(f"ENVIAR CERTIFICADOS - user: {user_email}, app_password: {bool(user_app_password)}")
    
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Evento no encontrado'}, status=404)
    
    certs = Certificate.objects.filter(event=event, status='generated').select_related('student')
    
    logger.info(f"Certificados con status=generated: {certs.count()}")
    
    if certs.count() == 0:
        return Response({
            'error': 'No hay certificados generados. Haz clic en "Generar certificados" primero.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not user_app_password:
        return Response({
            'error': 'No tienes App Password configurado. Ve a Mi Perfil y configúralo.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    sent = 0
    errors = []
    
    for cert in certs:
        logger.info(f"Procesando: {cert.student.email}")
        
        try:
            connection = EmailBackend(
                host='smtp.gmail.com',
                port=587,
                username=user_email,
                password=user_app_password,
                use_tls=True,
                fail_silently=False
            )
            
            subject = f"🎓 Certificado: {event.name}"
            message = f"""
¡Felicidades, {cert.student.first_name}!

Tu certificado del evento está listo. Encuentra el PDF adjunto en este email.

📌 {event.name}
📅 {event.event_date.strftime('%d de %B de %Y')}

¡Sigue adelante!

--
CertifyPro
            """.strip()
            
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=user_email,
                to=[cert.student.email],
                connection=connection
            )
            
            # Adjuntar PDF si existe
            if cert.pdf_file and cert.pdf_file.path:
                try:
                    email.attach_file(cert.pdf_file.path)
                    logger.info(f"PDF adjunto: {cert.pdf_file.path}")
                except Exception as pdf_error:
                    logger.error(f"Error adjuntando PDF: {pdf_error}")
            
            email.send(fail_silently=False)
            connection.close()
            sent += 1
            logger.info(f"✓ Email enviado a {cert.student.email}")
            
        except SMTPException as e:
            error_msg = str(e)
            logger.error(f"Error SMTP: {error_msg}")
            errors.append(f'{cert.student.email}: {error_msg}')
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error: {error_msg}")
            errors.append(f'{cert.student.email}: {error_msg}')
    
    logger.info(f"RESULTADO: enviados={sent}, errores={len(errors)}")
    
    if sent == 0:
        return Response({
            'success': False,
            'sent': 0,
            'message': errors[0] if errors else 'Error al enviar',
            'errors': errors[:10]
        }, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({
        'success': True,
        'sent': sent,
        'message': f'{sent} certificado(s) enviado(s) por email con PDF adjunto',
        'errors': errors[:10]
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def generar_certificados_aceptados(request, event_id):
    """
    Genera los certificados de invitados que aceptaron la invitación.
    Crea PDFs reales con la plantilla del evento.
    POST /api/events/{event_id}/generar-aceptados/
    """
    from django.conf import settings
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch, cm
    from reportlab.lib.utils import ImageReader
    import io
    import os
    from django.core.files.base import ContentFile
    
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Evento no encontrado'}, status=404)
    
    logger.info(f"=== GENERAR CERTIFICADOS ===")
    logger.info(f"Evento: {event.name} (ID: {event.id})")
    logger.info(f"Template image: {event.template_image}")
    logger.info(f"Name X: {event.name_x}, Name Y: {event.name_y}, Font: {event.name_font_size}")
    
    if not event.template_image:
        logger.error("No hay plantilla de imagen subida")
        return Response({
            'error': 'Primero sube una plantilla de imagen (JPG/PNG) para el certificado'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Solo certificados de estudiantes que aceptaron la invitación
    certs = Certificate.objects.filter(
        event=event,
        status='pending',
        student__enrollments__status='confirmed'
    ).select_related('student').distinct()
    
    logger.info(f"Certificados pendientes con enrollment confirmado: {certs.count()}")
    
    if certs.count() == 0:
        # Mostrar info de diagnóstico
        all_certs = Certificate.objects.filter(event=event)
        logger.info(f"Total certificados en evento: {all_certs.count()}")
        
        pending_certs = Certificate.objects.filter(event=event, status='pending')
        logger.info(f"Certificados pending: {pending_certs.count()}")
        
        enrollments = Enrollment.objects.filter(event=event, status='confirmed')
        logger.info(f"Enrollments confirmados: {enrollments.count()}")
        
        return Response({
            'error': 'No hay invitados que hayan aceptado la invitación. Un estudiante debe aceptar la invitación primero.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    generated = 0
    errors = []
    
    # Obtener la ruta de la imagen
    try:
        template_path = event.template_image.path
        logger.info(f"Template path: {template_path}")
    except Exception as e:
        logger.error(f"Error obteniendo path de imagen: {e}")
        return Response({
            'error': 'Error al acceder a la plantilla. Sube la imagen nuevamente.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Crear directorio para certificados si no existe
    media_root = settings.MEDIA_ROOT
    cert_dir = os.path.join(media_root, 'certificates', str(event.id))
    os.makedirs(cert_dir, exist_ok=True)
    logger.info(f"Directorio certificados: {cert_dir}")
    
    for cert in certs:
        try:
            student_name = f"{cert.student.first_name} {cert.student.last_name}".strip()
            if not student_name:
                student_name = cert.student.email.split('@')[0]
            
            logger.info(f"Generando certificado para: {student_name}")
            
            # Abrir la imagen plantilla
            img = Image.open(template_path)
            
            # Convertir a RGB si es necesario
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            # Obtener dimensiones de la imagen
            img_width, img_height = img.size
            
            # Crear objeto para dibujar
            draw = ImageDraw.Draw(img)
            
            # Cargar fuente
            font_size = event.name_font_size or 24
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
                except Exception:
                    font = ImageFont.load_default()
                    font_size = 20
            
            # Obtener tamaño del texto
            try:
                bbox = draw.textbbox((0, 0), student_name, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except Exception:
                text_width = font_size * len(student_name)
                text_height = font_size
            
            # Posición X e Y desde el evento
            x = event.name_x if event.name_x else (img_width // 2)
            y = event.name_y if event.name_y else (img_height // 2)
            
            # El lado DERECHO del texto está en la posición X
            text_x = x - text_width
            text_y = y - (text_height // 2)
            
            # Dibujar texto con sombra
            shadow_color = (200, 200, 200)
            draw.text((text_x + 1, text_y + 1), student_name, font=font, fill=shadow_color)
            draw.text((text_x, text_y), student_name, font=font, fill=(0, 0, 0))
            
            # Guardar como PNG temporal
            temp_png = io.BytesIO()
            img.save(temp_png, format='PNG')
            temp_png.seek(0)
            
            # Crear PDF con reportlab
            pdf_buffer = io.BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=(img_width, img_height))
            c.drawImage(ImageReader(temp_png), 0, 0, width=img_width, height=img_height)
            c.save()
            pdf_buffer.seek(0)
            
            # Guardar el PDF en el certificado
            pdf_filename = f"certificado_{cert.id}_{student_name.replace(' ', '_')}.pdf"
            cert.pdf_file.save(pdf_filename, ContentFile(pdf_buffer.read()), save=False)
            
            # Actualizar status
            cert.status = 'generated'
            cert.save(update_fields=['status', 'pdf_file'])
            
            generated += 1
            logger.info(f"✓ Certificado generado: {student_name}")
            
        except Exception as e:
            import traceback
            error_msg = str(e)
            logger.error(f"Error generando certificado para {cert.student.email}: {error_msg}")
            logger.error(traceback.format_exc())
            errors.append(f'{cert.student.email}: {error_msg}')
    
    logger.info(f"Generados: {generated}, Errores: {len(errors)}")
    
    return Response({
        'success': True,
        'generated': generated,
        'message': f'{generated} certificado(s) generado(s) con PDF' if generated > 0 else 'No hay invitados que hayan aceptado',
        'errors': errors[:10]
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def listar_certificados_evento(request, event_id):
    """
    Lista los certificados de un evento con sus estados.
    GET /api/events/{event_id}/certificados/
    """
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Evento no encontrado'}, status=404)
    
    # Solo certificados de estudiantes que aceptaron (enrollment confirmado)
    certs = Certificate.objects.filter(
        event=event,
        student__enrollments__status='confirmed'
    ).select_related('student').distinct()
    
    data = [{
        'id': c.id,
        'student_name': f"{c.student.first_name} {c.student.last_name}".strip(),
        'student_email': c.student.email,
        'student_phone': c.student.phone or '',
        'status': c.status,
    } for c in certs]
    
    stats = {
        'total': certs.count(),
        'pending': certs.filter(status='pending').count(),
        'generated': certs.filter(status='generated').count(),
    }
    
    return Response({
        'certificates': data,
        'stats': stats
    })